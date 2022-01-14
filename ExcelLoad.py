import os
import openpyxl
FilePath = os.getcwd()+"\Resources\Data.xlsx"
Sheet = openpyxl.load_workbook(FilePath)
ActiveSheet = Sheet.active
columnCount = ActiveSheet.max_column
rowCount = ActiveSheet.max_row

def TestData(Scenario , columnName):
    for i in range(1, rowCount+1):
        column_name = ActiveSheet.cell(row = i, column = 1)
        if column_name.value == Scenario:
            break
    for k in range(1, columnCount+1):
        row_value = ActiveSheet.cell(row = 1, column=k)
        if row_value.value == columnName:
            break
    value = ActiveSheet.cell(row = i, column = k).value
    return value
